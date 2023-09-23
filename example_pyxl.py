import openpyxl
import pd
from openpyxl.chart import BarChart, Reference

def process_xls(filename):
    wb = openpyxl.load_workbook(filename)
    wb2 = pd.read_excel(filename)
    sheet = wb['Sheet1']
    sheet2 = wb2['Sheet1']
    cell = sheet['a1'] #sheet.cell(1,1)
    print(cell.value)
    print(sheet.max_row)
    print(sheet2.cell)

    for row in range(2,sheet.max_row +1):
        cell = sheet.cell(row,3)
        corrected_value = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_value

    values = Reference(sheet, 
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')
    wb.save('transaction2.xlsx')

process_xls('transaction.xlsx')
